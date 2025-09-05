import os
from datetime import datetime, timedelta
from typing import Any, Dict, List

from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.db.models import Sum, Count
from django.utils import timezone

from .models import (
    KV, Daily, UserProfile,
    VoiceUserDaily, VoiceUserTotal,
    MessageUserDaily, MessageUserTotal,
    VoiceChannel, VoiceChannelDaily, VoiceUserChannelDaily,
)

# ================== helpers ==================

BACKDATE_DAYS = int(os.getenv("BACKDATE_DAYS", "0"))  

def _logic_date():
    return timezone.localdate() - timedelta(days=BACKDATE_DAYS)

def _get_total_messages() -> int:
    try:
        return int(KV.objects.get(pk="messages_total").val)
    except KV.DoesNotExist:
        KV.objects.create(key="messages_total", val="0")
        return 0

def _profile_map(user_ids: List[str]) -> Dict[str, Dict[str, Any]]:
    if not user_ids:
        return {}
    rows = UserProfile.objects.filter(user_id__in=user_ids).values(
        "user_id", "username", "display_name", "avatar_url", "joined_at"
    )
    d: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        d[r["user_id"]] = {
            "username": r["username"] or "",
            "display_name": r["display_name"] or "",
            "avatar_url": r["avatar_url"] or "https://cdn.discordapp.com/embed/avatars/0.png",
            "joined_at": r.get("joined_at"),
        }
    return d

def _excel_safe(v: Any):
    if isinstance(v, datetime):
        return v.replace(tzinfo=None)
    try:
        return v.isoformat()  # type: ignore[attr-defined]
    except Exception:
        return v

# ================== NOW / HISTORY ==================

def now(request):
    d = _logic_date()
    row, _ = Daily.objects.get_or_create(
        date=d,
        defaults={
            "members": 0, "joins": 0, "leaves": 0,
            "messages": 0, "messages_total": _get_total_messages(),
            "voice_seconds": 0,
        },
    )

    active_authors = MessageUserDaily.objects.filter(date=d).values("user_id").distinct().count()
    active_voice   = VoiceUserDaily.objects.filter(date=d).values("user_id").distinct().count()
    visitors       = max(active_authors, active_voice)
    msgs_today     = int(row.messages or 0)
    avg_per_active = round(msgs_today / active_authors, 2) if active_authors else 0.0

    return JsonResponse({
        "date": str(d),
        "members": int(row.members or 0),
        "joins": int(row.joins or 0),
        "leaves": int(row.leaves or 0),
        "diff": int(row.joins or 0) - int(row.leaves or 0),
        "messages_today": msgs_today,
        "messages_total": int(row.messages_total or 0),
        "voice_hours_today": round((int(row.voice_seconds or 0)) / 3600, 2),
        "unique_message_members": active_authors,
        "visitors": visitors,
        "avg_messages_per_active_member": avg_per_active,
    })

def history(request):
    rows = list(
        Daily.objects.order_by("-date").values(
            "date", "members", "joins", "leaves",
            "messages", "messages_total", "voice_seconds"
        )
    )
    for r in rows:
        r["voice_hours"] = round((int(r.get("voice_seconds") or 0)) / 3600, 2)
    return JsonResponse(rows, safe=False)

# ================== VOICE (LISTS) ==================

def voice_today(request):
    d = _logic_date()
    rows = list(
        VoiceUserDaily.objects.filter(date=d)
        .order_by("-seconds")
        .values("user_id", "seconds")
    )
    prof = _profile_map([r["user_id"] for r in rows])
    out: List[Dict[str, Any]] = []
    for r in rows:
        p = prof.get(r["user_id"], {})
        sec = int(r["seconds"] or 0)
        out.append({
            "user_id": r["user_id"],
            "username": p.get("username", ""),
            "display_name": p.get("display_name", ""),
            "avatar_url": p.get("avatar_url"),
            "seconds": sec,
            "hours": round(sec / 3600, 2),
        })
    return JsonResponse(out, safe=False)

def voice_by_date(request):
    q = request.GET.get("date")
    if not q:
        return HttpResponseBadRequest("date required YYYY-MM-DD")
    rows = list(
        VoiceUserDaily.objects.filter(date=q)
        .order_by("-seconds")
        .values("user_id", "seconds")
    )
    prof = _profile_map([r["user_id"] for r in rows])
    out: List[Dict[str, Any]] = []
    for r in rows:
        p = prof.get(r["user_id"], {})
        sec = int(r["seconds"] or 0)
        out.append({
            "user_id": r["user_id"],
            "username": p.get("username", ""),
            "display_name": p.get("display_name", ""),
            "avatar_url": p.get("avatar_url"),
            "seconds": sec,
            "hours": round(sec / 3600, 2),
        })
    return JsonResponse(out, safe=False)

# ===== VOICE BY CHANNEL (LISTS) =====

def voice_channels_today(request):
    d = _logic_date()
    rows = list(
        VoiceChannelDaily.objects.filter(date=d)
        .order_by("-seconds")
        .values("channel_id", "seconds")
    )
    id2name = dict(
        VoiceChannel.objects.filter(channel_id__in=[r["channel_id"] for r in rows])
        .values_list("channel_id", "name")
    )
    out = [{
        "channel_id": r["channel_id"],
        "channel_name": id2name.get(r["channel_id"], ""),
        "seconds": int(r["seconds"] or 0),
        "hours": round((int(r["seconds"] or 0)) / 3600, 2),
    } for r in rows]
    return JsonResponse(out, safe=False)

def voice_channel_users_today(request, channel_id: str):
    d = _logic_date()
    rows = list(
        VoiceUserChannelDaily.objects.filter(date=d, channel_id=channel_id)
        .order_by("-seconds")
        .values("user_id", "seconds")
    )
    prof = _profile_map([r["user_id"] for r in rows])
    out: List[Dict[str, Any]] = []
    for r in rows:
        p = prof.get(r["user_id"], {})
        sec = int(r["seconds"] or 0)
        out.append({
            "user_id": r["user_id"],
            "username": p.get("username", ""),
            "display_name": p.get("display_name", ""),
            "avatar_url": p.get("avatar_url"),
            "seconds": sec,
            "hours": round(sec / 3600, 2),
        })
    return JsonResponse(out, safe=False)

# ================== VOICE (BY USER) ==================

def voice_user_today(request, user_id: str):
    d = _logic_date()
    sec = VoiceUserDaily.objects.filter(date=d, user_id=user_id)\
        .values_list("seconds", flat=True).first() or 0
    prof = _profile_map([user_id]).get(user_id, {"user_id": user_id})
    return JsonResponse({
        "user": prof,
        "seconds": int(sec),
        "hours": round(int(sec) / 3600, 2),
    })

def voice_user_history(request, user_id: str):
    rows = list(
        VoiceUserDaily.objects.filter(user_id=user_id)
        .order_by("-date")
        .values("date", "seconds")
    )
    out = [{
        "date": str(r["date"]),
        "seconds": int(r["seconds"] or 0),
        "hours": round((int(r["seconds"] or 0)) / 3600, 2)
    } for r in rows]
    return JsonResponse(out, safe=False)

def voice_user_total(request, user_id: str):
    tot = VoiceUserTotal.objects.filter(user_id=user_id)\
        .values_list("seconds", flat=True).first() or 0
    return JsonResponse({
        "user_id": user_id,
        "seconds": int(tot),
        "hours": round(int(tot) / 3600, 2),
    })

# ================== MESSAGES ==================

def messages_users_today(request):
    d = _logic_date()
    rows = list(
        MessageUserDaily.objects.filter(date=d)
        .order_by("-messages")
        .values("user_id", "messages")
    )
    prof = _profile_map([r["user_id"] for r in rows])
    out: List[Dict[str, Any]] = []
    for r in rows:
        p = prof.get(r["user_id"], {})
        out.append({
            "user_id": r["user_id"],
            "username": p.get("username", ""),
            "display_name": p.get("display_name", ""),
            "avatar_url": p.get("avatar_url"),
            "messages": int(r["messages"] or 0),
        })
    return JsonResponse(out, safe=False)

def messages_user_today(request, user_id: str):
    d = _logic_date()
    cnt = MessageUserDaily.objects.filter(date=d, user_id=user_id)\
        .values_list("messages", flat=True).first() or 0
    prof = _profile_map([user_id]).get(user_id, {"user_id": user_id})
    return JsonResponse({"user": prof, "messages": int(cnt)})

def messages_user_history(request, user_id: str):
    rows = MessageUserDaily.objects.filter(user_id=user_id)\
        .order_by("-date").values("date", "messages")
    out = [{"date": str(r["date"]), "messages": int(r["messages"] or 0)} for r in rows]
    return JsonResponse(out, safe=False)

def messages_user_total(request, user_id: str):
    total = MessageUserTotal.objects.filter(user_id=user_id)\
        .values_list("messages", flat=True).first()
    if total is None:
        total = MessageUserDaily.objects.filter(user_id=user_id)\
            .aggregate(s=Sum("messages"))["s"] or 0
    return JsonResponse({"user_id": user_id, "messages": int(total)})

# ================== USER (summary today) ==================

def user_today(request, user_id: str):
    d = _logic_date()
    prof = _profile_map([user_id]).get(user_id, {"user_id": user_id})
    voice_sec = VoiceUserDaily.objects.filter(date=d, user_id=user_id)\
        .values_list("seconds", flat=True).first() or 0
    msg_cnt = MessageUserDaily.objects.filter(date=d, user_id=user_id)\
        .values_list("messages", flat=True).first() or 0
    return JsonResponse({
        "user": prof,
        "voice_seconds": int(voice_sec),
        "voice_hours": round(int(voice_sec) / 3600, 2),
        "messages": int(msg_cnt),
    })

# ================== EXPORT (XLSX) ==================

def export_xlsx(request):
    """
    Полный исторический экспорт.
    Листы:
      - Daily (+ unique_message_members, avg_messages_per_active_member)
      - VoiceByChannelDay
      - VoiceUserByChannelDay
      - MessagesByDay
      - Profiles
    """
    from openpyxl import Workbook
    import io

    wb = Workbook()

    # Daily
    ws = wb.active
    ws.title = "Daily"
    ws.append([
        "date", "members", "joins", "leaves",
        "messages", "messages_total", "voice_seconds", "voice_hours",
        "unique_message_members", "avg_messages_per_active_member",
    ])
    authors_per_day = dict(
        MessageUserDaily.objects
        .values("date")
        .annotate(c=Count("user_id", distinct=True))
        .values_list("date", "c")
    )
    for r in Daily.objects.order_by("date").values(
        "date", "members", "joins", "leaves", "messages", "messages_total", "voice_seconds"
    ):
        msgs = int(r["messages"] or 0)
        authors = int(authors_per_day.get(r["date"], 0) or 0)
        avg = round(msgs / authors, 2) if authors else 0.0
        ws.append([
            _excel_safe(r["date"]),
            r["members"] or 0, r["joins"] or 0, r["leaves"] or 0,
            msgs, r["messages_total"] or 0,
            r["voice_seconds"] or 0, round((int(r["voice_seconds"] or 0)) / 3600, 2),
            authors, avg
        ])

    # VoiceByChannelDay
    ws2 = wb.create_sheet("VoiceByChannelDay")
    ws2.append(["date", "channel_id", "channel_name", "seconds", "hours"])
    id2name = dict(VoiceChannel.objects.values_list("channel_id", "name"))
    for r in VoiceChannelDaily.objects.order_by("date", "channel_id").values("date", "channel_id", "seconds"):
        sec = int(r["seconds"] or 0)
        ws2.append([
            _excel_safe(r["date"]),
            r["channel_id"], id2name.get(r["channel_id"], ""),
            sec, round(sec / 3600, 2)
        ])

    # VoiceUserByChannelDay
    ws3 = wb.create_sheet("VoiceUserByChannelDay")
    ws3.append(["date", "channel_id", "channel_name", "user_id", "seconds", "hours"])
    for r in VoiceUserChannelDaily.objects.order_by("date", "channel_id", "user_id")\
            .values("date", "channel_id", "user_id", "seconds"):
        sec = int(r["seconds"] or 0)
        ws3.append([
            _excel_safe(r["date"]),
            r["channel_id"], id2name.get(r["channel_id"], ""),
            r["user_id"], sec, round(sec / 3600, 2)
        ])

    # MessagesByDay
    ws4 = wb.create_sheet("MessagesByDay")
    ws4.append(["user_id", "date", "messages"])
    for r in MessageUserDaily.objects.order_by("date", "user_id").values("user_id", "date", "messages"):
        ws4.append([r["user_id"], _excel_safe(r["date"]), int(r["messages"] or 0)])

    # Profiles
    ws5 = wb.create_sheet("Profiles")
    ws5.append(["user_id", "username", "display_name", "avatar_url", "joined_at", "is_bot"])
    for p in UserProfile.objects.order_by("user_id").values(
        "user_id", "username", "display_name", "avatar_url", "joined_at", "is_bot"
    ):
        ws5.append([
            p["user_id"], p.get("username") or "", p.get("display_name") or "",
            p.get("avatar_url") or "", _excel_safe(p.get("joined_at") or ""), bool(p.get("is_bot"))
        ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="discord_metrics_all.xlsx"'
    return resp